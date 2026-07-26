import os
import io
import csv
import calendar
import logging
import sqlite3
import shutil
from typing import List, Optional, Any
from datetime import datetime, timedelta
from contextlib import asynccontextmanager
from dotenv import load_dotenv

load_dotenv()
from fastapi import FastAPI, Depends, HTTPException, Query, Header, Request, status, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, ConfigDict
from sqlalchemy.orm import Session

from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

from database import init_db, get_db, SessionLocal
from models import Transaction, Category, Budget, RecurringRule

# Report Generation Libraries
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

logging.basicConfig(level=logging.INFO)
limiter = Limiter(key_func=get_remote_address)


def require_api_key(x_api_key: Optional[str] = Header(None, alias="X-API-Key")):
    api_key_env = os.getenv("API_KEY", "").strip()
    if api_key_env:
        if not x_api_key or x_api_key != api_key_env:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid or missing X-API-Key header"
            )


@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    api_key_env = os.getenv("API_KEY", "").strip()
    if not api_key_env:
        logging.warning("API running without authentication")
    yield


app = FastAPI(
    title="Personal Finance Tracker API",
    description="REST API for Personal Finance Tracker V2",
    version="2.0.0",
    lifespan=lifespan,
)

app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

cors_env = os.getenv("CORS_ORIGINS", "http://localhost:8000,http://127.0.0.1:8000")
origins = [o.strip() for o in cors_env.split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------------
# Pydantic Schemas
# ---------------------------------------------------------------------------
class TransactionSchema(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: Optional[int] = None
    date: str
    amount: float
    category: str
    description: Optional[str] = ""
    currency: Optional[str] = "INR"


class CategorySchema(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: Optional[int] = None
    name: str
    type: str  # "Income" or "Expense"
    color: Optional[str] = "#60a5fa"
    icon: Optional[str] = "📋"


class BudgetSchema(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: Optional[int] = None
    category_name: str
    monthly_limit: float


class RecurringSchema(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: Optional[int] = None
    description: str
    amount: float
    category: str
    frequency: str = "monthly"  # "monthly" or "weekly"
    next_date: str
    is_active: bool = True


# ---------------------------------------------------------------------------
# Root Route: Serve Dashboard
# ---------------------------------------------------------------------------
@app.get("/", response_class=HTMLResponse)
def read_root():
    if os.path.exists("index.html"):
        with open("index.html", "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read())
    return HTMLResponse("<h1>Personal Finance Tracker V2 API Running</h1>")


# ---------------------------------------------------------------------------
# Transactions Endpoints
# ---------------------------------------------------------------------------
@app.get("/api/transactions", response_model=List[TransactionSchema])
def get_transactions(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    category: Optional[str] = None,
    type: Optional[str] = None,
    db: Session = Depends(get_db),
):
    query = db.query(Transaction)
    if category:
        query = query.filter(Transaction.category == category)
    if type:
        income_cats = [c.name for c in db.query(Category).filter(Category.type == "Income").all()]
        if type == "Income":
            query = query.filter(Transaction.category.in_(income_cats))
        elif type == "Expense":
            query = query.filter(~Transaction.category.in_(income_cats))

    transactions = query.all()

    if start_date or end_date:
        fmt = "%d-%m-%Y"
        filtered = []
        for t in transactions:
            try:
                t_date = datetime.strptime(t.date, fmt)
                if start_date and t_date < datetime.strptime(start_date, fmt):
                    continue
                if end_date and t_date > datetime.strptime(end_date, fmt):
                    continue
                filtered.append(t)
            except ValueError:
                filtered.append(t)
        transactions = filtered

    # Sort descending by date
    try:
        transactions.sort(
            key=lambda x: datetime.strptime(x.date, "%d-%m-%Y"), reverse=True
        )
    except Exception:
        pass

    return transactions


@app.post("/api/transactions", response_model=TransactionSchema, dependencies=[Depends(require_api_key)])
def create_transaction(item: TransactionSchema, db: Session = Depends(get_db)):
    tx = Transaction(
        date=item.date,
        amount=item.amount,
        category=item.category,
        description=item.description or "",
        currency=item.currency or "INR",
    )
    db.add(tx)
    db.commit()
    db.refresh(tx)
    return tx


@app.put("/api/transactions/{tx_id}", response_model=TransactionSchema, dependencies=[Depends(require_api_key)])
def update_transaction(
    tx_id: int, item: TransactionSchema, db: Session = Depends(get_db)
):
    tx = db.query(Transaction).filter(Transaction.id == tx_id).first()
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    tx.date = item.date
    tx.amount = item.amount
    tx.category = item.category
    tx.description = item.description or ""
    if item.currency:
        tx.currency = item.currency
    db.commit()
    db.refresh(tx)
    return tx


@app.delete("/api/transactions/{tx_id}", dependencies=[Depends(require_api_key)])
def delete_transaction(tx_id: int, db: Session = Depends(get_db)):
    tx = db.query(Transaction).filter(Transaction.id == tx_id).first()
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    db.delete(tx)
    db.commit()
    return {"message": "Transaction deleted successfully"}


# ---------------------------------------------------------------------------
# Categories Endpoints
# ---------------------------------------------------------------------------
@app.get("/api/categories", response_model=List[CategorySchema])
def get_categories(db: Session = Depends(get_db)):
    return db.query(Category).all()


@app.post("/api/categories", response_model=CategorySchema, dependencies=[Depends(require_api_key)])
def create_category(item: CategorySchema, db: Session = Depends(get_db)):
    existing = db.query(Category).filter(Category.name == item.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Category already exists")
    cat = Category(
        name=item.name,
        type=item.type,
        color=item.color or "#60a5fa",
        icon=item.icon or "📋",
    )
    db.add(cat)
    db.commit()
    db.refresh(cat)
    return cat


@app.delete("/api/categories/{cat_id}", dependencies=[Depends(require_api_key)])
def delete_category(cat_id: int, db: Session = Depends(get_db)):
    cat = db.query(Category).filter(Category.id == cat_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    
    # Check referential integrity across dependent tables
    has_tx = db.query(Transaction).filter(Transaction.category == cat.name).first()
    has_budget = db.query(Budget).filter(Budget.category_name == cat.name).first()
    has_recurring = db.query(RecurringRule).filter(RecurringRule.category == cat.name).first()

    if has_tx or has_budget or has_recurring:
        referenced_in = []
        if has_tx:
            referenced_in.append("transactions")
        if has_budget:
            referenced_in.append("budgets")
        if has_recurring:
            referenced_in.append("recurring rules")
        refs_str = ", ".join(referenced_in)
        raise HTTPException(
            status_code=409,
            detail=f"Cannot delete category '{cat.name}' because it is referenced by existing {refs_str}."
        )

    db.delete(cat)
    db.commit()
    return {"message": "Category deleted successfully"}


# ---------------------------------------------------------------------------
# Budgets Endpoints
# ---------------------------------------------------------------------------
@app.get("/api/budgets", response_model=List[BudgetSchema])
def get_budgets(db: Session = Depends(get_db)):
    return db.query(Budget).all()


@app.post("/api/budgets", response_model=BudgetSchema, dependencies=[Depends(require_api_key)])
def set_budget(item: BudgetSchema, db: Session = Depends(get_db)):
    existing = (
        db.query(Budget)
        .filter(Budget.category_name == item.category_name)
        .first()
    )
    if existing:
        existing.monthly_limit = item.monthly_limit
        db.commit()
        db.refresh(existing)
        return existing
    b = Budget(category_name=item.category_name, monthly_limit=item.monthly_limit)
    db.add(b)
    db.commit()
    db.refresh(b)
    return b


@app.delete("/api/budgets/{category_name}", dependencies=[Depends(require_api_key)])
def delete_budget(category_name: str, db: Session = Depends(get_db)):
    b = db.query(Budget).filter(Budget.category_name == category_name).first()
    if not b:
        raise HTTPException(status_code=404, detail="Budget not found")
    db.delete(b)
    db.commit()
    return {"message": "Budget removed successfully"}


# ---------------------------------------------------------------------------
# Recurring Rules Endpoints
# ---------------------------------------------------------------------------
@app.get("/api/recurring", response_model=List[RecurringSchema])
def get_recurring(db: Session = Depends(get_db)):
    return db.query(RecurringRule).all()


@app.post("/api/recurring", response_model=RecurringSchema, dependencies=[Depends(require_api_key)])
def create_recurring(item: RecurringSchema, db: Session = Depends(get_db)):
    rule = RecurringRule(
        description=item.description,
        amount=item.amount,
        category=item.category,
        frequency=item.frequency or "monthly",
        next_date=item.next_date,
        is_active=item.is_active,
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return rule


@app.delete("/api/recurring/{rule_id}", dependencies=[Depends(require_api_key)])
def delete_recurring(rule_id: int, db: Session = Depends(get_db)):
    rule = db.query(RecurringRule).filter(RecurringRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Recurring rule not found")
    db.delete(rule)
    db.commit()
    return {"message": "Recurring rule deleted"}


@app.post("/api/recurring/process", dependencies=[Depends(require_api_key)])
def process_recurring(db: Session = Depends(get_db)):
    today_str = datetime.today().strftime("%d-%m-%Y")
    today_dt = datetime.strptime(today_str, "%d-%m-%Y")

    rules = db.query(RecurringRule).filter(RecurringRule.is_active == True).all()
    created_txs = []

    for rule in rules:
        try:
            next_dt = datetime.strptime(rule.next_date, "%d-%m-%Y")
            orig_day = next_dt.day
            while next_dt <= today_dt:
                # Add transaction
                tx = Transaction(
                    date=next_dt.strftime("%d-%m-%Y"),
                    amount=rule.amount,
                    category=rule.category,
                    description=f"[Recurring] {rule.description}",
                    currency="INR",
                )
                db.add(tx)
                created_txs.append(rule.description)

                # Increment next date
                if rule.frequency == "weekly":
                    next_dt += timedelta(days=7)
                else:  # monthly
                    month = next_dt.month % 12 + 1
                    year = next_dt.year + (next_dt.month // 12)
                    max_days = calendar.monthrange(year, month)[1]
                    day = min(orig_day, max_days)
                    next_dt = datetime(year, month, day)

            rule.next_date = next_dt.strftime("%d-%m-%Y")
        except Exception as e:
            print(f"Error processing recurring rule {rule.id}: {e}")

    db.commit()
    return {"processed_count": len(created_txs), "items": created_txs}


# ---------------------------------------------------------------------------
# Export Endpoints (PDF, Excel, CSV)
# ---------------------------------------------------------------------------
@app.get("/api/export/excel")
@limiter.limit("10/minute")
def export_excel(request: Request, db: Session = Depends(get_db)):
    transactions = db.query(Transaction).all()

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Transactions Report"

    # Header styles
    header_fill = PatternFill(
        start_color="1E293B", end_color="1E293B", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    headers = ["ID", "Date", "Category", "Amount (₹)", "Description", "Currency"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Data Rows
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for tx in transactions:
        row = [tx.id, tx.date, tx.category, tx.amount, tx.description, tx.currency]
        ws.append(row)
        current_row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            if col == 4:
                cell.number_format = "₹#,##0.00"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Financial_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(
        output,
        headers=headers,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/export/pdf")
@limiter.limit("10/minute")
def export_pdf(request: Request, db: Session = Depends(get_db)):
    transactions = db.query(Transaction).all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    elements: list[Any] = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#0F172A"),
        spaceAfter=12,
    )

    elements.append(Paragraph("💰 Personal Finance Summary Report", title_style))
    elements.append(
        Paragraph(
            f"Generated on: {datetime.now().strftime('%d %B %Y, %H:%M')}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 16))

    # Table Header
    data = [["Date", "Category", "Amount (₹)", "Description"]]

    total_inc = 0
    total_exp = 0

    income_cats = [c.name for c in db.query(Category).filter(Category.type == "Income").all()]

    for tx in transactions:
        if tx.category in income_cats:
            total_inc += tx.amount
        else:
            total_exp += tx.amount
        data.append([tx.date, tx.category, f"₹{tx.amount:,.2f}", tx.description or "-"])

    t = Table(data, colWidths=[80, 100, 100, 240])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ]
        )
    )

    elements.append(t)
    elements.append(Spacer(1, 20))

    # Summary Section
    summary_text = (
        f"<b>Total Income:</b> ₹{total_inc:,.2f} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Total Expenses:</b> ₹{total_exp:,.2f} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Net Savings:</b> ₹{total_inc - total_exp:,.2f}"
    )
    elements.append(Paragraph(summary_text, styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)

    filename = f"Financial_Summary_{datetime.now().strftime('%Y%m%d')}.pdf"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(
        buffer, headers=headers, media_type="application/pdf"
    )


@app.get("/api/export/csv")
def export_csv_api(db: Session = Depends(get_db)):
    transactions = db.query(Transaction).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["date", "amount", "category", "description", "currency"])

    for tx in transactions:
        writer.writerow([tx.date, tx.amount, tx.category, tx.description, tx.currency])

    mem = io.BytesIO()
    mem.write(output.getvalue().encode("utf-8"))
    mem.seek(0)

    filename = f"finance_data_export_{datetime.now().strftime('%Y%m%d')}.csv"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(
        mem, headers=headers, media_type="text/csv"
    )


# ---------------------------------------------------------------------------
# Backup & Restore Endpoints
# ---------------------------------------------------------------------------
@app.get("/api/backup")
def backup_database():
    db_url = os.getenv("DATABASE_URL", "sqlite:///./finance.db")
    db_file = db_url.replace("sqlite:///", "")
    if not os.path.exists(db_file):
        raise HTTPException(status_code=404, detail="Database file not found")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return FileResponse(
        db_file,
        filename=f"finance_backup_{timestamp}.db",
        media_type="application/x-sqlite3",
    )


@app.post("/api/restore", dependencies=[Depends(require_api_key)])
async def restore_database(file: UploadFile = File(...)):
    if not file.filename or not file.filename.endswith((".db", ".sqlite", ".sqlite3")):
        raise HTTPException(status_code=400, detail="Invalid file type. Must be a .db or .sqlite file.")

    temp_path = "temp_restore.db"
    try:
        content = await file.read()
        with open(temp_path, "wb") as f:
            f.write(content)

        # Validate SQLite format and required schema tables
        conn = sqlite3.connect(temp_path)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = {row[0] for row in cursor.fetchall()}
        conn.close()

        required_tables = {"transactions", "categories", "budgets", "recurring_rules"}
        if not required_tables.issubset(tables):
            raise HTTPException(
                status_code=400,
                detail=f"Invalid database schema. Missing required tables: {required_tables - tables}",
            )

        db_url = os.getenv("DATABASE_URL", "sqlite:///./finance.db")
        target_db = db_url.replace("sqlite:///", "")
        if os.path.exists(target_db):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            bak_path = f"{target_db}.bak-{timestamp}"
            shutil.copy2(target_db, bak_path)

        shutil.move(temp_path, target_db)
        return {"message": "Database restored successfully"}
    except HTTPException:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise HTTPException(status_code=400, detail=f"Failed to restore database: {str(e)}")
